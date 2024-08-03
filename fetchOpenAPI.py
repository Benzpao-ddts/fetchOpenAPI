import requests
from openpyxl import Workbook
import os
import re

# URL to fetch the OpenAPI specification
openapi_url = "https://petstore.swagger.io/v2/swagger.json"

# Fetch the OpenAPI specification
response = requests.get(openapi_url)
openapi_spec = response.json()

# Configuration settings
use_keyword = True          # Set to False to extract all paths
keyword = "pet"             # Keyword to filter paths
pattern = re.compile(keyword) if use_keyword else None

# Create a new Excel workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Insert the headers
ws['A1'] = "Path"
ws['B1'] = "Method"

# Initialize a row counter
row = 2

# Extract paths and methods from the OpenAPI specification
for path, methods in openapi_spec.get('paths', {}).items():
    if pattern:
        if pattern.search(path):  # Check if path matches the regex pattern
            for method in methods.keys():
                ws[f'A{row}'] = path
                ws[f'B{row}'] = method.upper()  # Convert method to uppercase for consistency
                row += 1
    else:
        for method in methods.keys():
            ws[f'A{row}'] = path
            ws[f'B{row}'] = method.upper()  # Convert method to uppercase for consistency
            row += 1

# Ensure the output directory exists
output_dir = "result"
os.makedirs(output_dir, exist_ok=True)

# Save the workbook in the project's folder under 'result' directory
output_file = os.path.join(output_dir, "output_with_path.xlsx")
wb.save(output_file)

print("Paths and methods extracted and saved into Excel file successfully.")
